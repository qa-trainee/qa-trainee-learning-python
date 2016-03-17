from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet()  # insert at the end (default)
ws2 = wb.create_sheet(0)  # insert at first position
ws.title = "New title"
ws.sheet_properties.tabColor = "1072BA"
ws3 = wb["New title"]
ws4 = wb.get_sheet_by_name("New title")
print(ws is ws3 is ws4)
print(wb.get_sheet_names())

for s in wb:
    print(sheet.title)


